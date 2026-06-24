import os
import re
import shutil
import subprocess
import pandas as pd
import openpyxl
import pdfplumber

# =====================================================
# PRODUCT MAPPINGS AND CONFIGS
# =====================================================

PRODUCT_MAP = {
    'MX150 2X10 BLD (33482)': 'MX150 BLD 2X10',
    'MX150 2X10 RCPT (33472)': 'MX150 RCPT 2X10',
    'CMX 32 CKT (34868)': 'CMX 32 CKT',
    'MX64 1X6 CKT (31403)': 'MX64 31403 1X6 CKT',
    'MX123 49 CKT WDC (216744)': 'MX123 49 CKT WDC',
    'MXDASH 15+1 RCPT (218342)': 'MXDASH 15+1 RCPT'
}

HOUSING_MAP = {
    "A-BK": "A-BLACK",
    "B-LG": "B-LIGHT GRAY",
    "C-DG": "C-DARK GRAY",
    "D-SG": "D-STONE GRAY",
    "A-YW": "A-YELLOW",
    "A-BK-CLIP": "A-BLACK-CLIP",
    "B-LG-CLIP": "B-LIGHT GRAY-CLIP",
    "C-DG-CLIP": "C-DARK GRAY-CLIP",
    "D-SG-CLIP": "D-STONE GRAY-CLIP",
    "A-YW-CLIP": "A-YELLOW-CLIP",
    "B-YW-CLIP": "B-YELLOW-CLIP",
    "C-": "C-BROWN",
    "D-": "D-GREEN"
}

PRODUCT_CATEGORIES = {
    'MX150 2X10 BLD (33482)': [
        'HOUSING KEY/COLOR',
        'GROMMET CAP',
        'PLR',
        'MAT SEAL'
    ],
    'MX150 2X10 RCPT (33472)': [
        'HOUSING KEY/COLOR',
        'GROMMET CAP',
        'PLR',
        'CPA',
        'MAT SEAL',
        'RING SEAL'
    ],
    'CMX 32 CKT (34868)': [
        'HOUSING',
        'TPA-KEY-COLOR',
        'R-SEAL',
        'RSC',
        'M-SEAL',
        'LEVER'
    ],
    'MX64 1X6 CKT (31403)': [
        'CON-HOUSING',
        'TERMINAL HOUSING',
        'TPA',
        'CPA',
        'MAT SEAL',
        'RING SEAL'
    ],
    'MX123 49 CKT WDC (216744)': [
        'CONNECTOR HOUSING',
        'GROMMET CAP',
        'TPA',
        'RIGHT SLIDE',
        'LEFT SLIDE',
        'MATTE ASSIST LEVER',
        'CPA',
        'MAT SEAL',
        'RING SEAL'
    ],
    'MXDASH 15+1 RCPT (218342)': [
        'CON-HOUSING',
        'TERMINAL HOUSING',
        'M-SEAL CAP',
        'ISL',
        'CPA',
        'MAT SEAL',
        'RING SEAL'
    ]
}

def clean_val(val):
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val.is_integer():
            val = int(val)
    val_str = str(val).strip()
    if val_str in ("", "-", "nan", "NaN", "None"):
        return ""
    return val_str

def format_ref_number(val):
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return int(val)
    try:
        return int(str(val).strip())
    except ValueError:
        return str(val).strip()

def parse_pdf_bom_data(pdf_path, p_name):
    pdf_data = {}
    if not os.path.exists(pdf_path):
        print(f"  Warning: PDF file not found: {pdf_path}")
        return pdf_data
        
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                lines = text.split("\n")
                for line in lines:
                    line_str = line.strip()
                    
                    # 1. Match Part/Reference Numbers based on product family
                    ref_number = None
                    
                    if p_name == 'MX150 2X10 BLD (33482)':
                        match = re.search(r"\b(33482\d{4})\b", line_str)
                        if match:
                            ref_number = match.group(1)
                    elif p_name == 'MX150 2X10 RCPT (33472)':
                        match = re.search(r"\b(3347\d{5,})\b", line_str)
                        if match:
                            ref_number = match.group(1)
                    elif p_name == 'CMX 32 CKT (34868)':
                        match = re.search(r"\b(34868\d{4})\b", line_str)
                        if match:
                            ref_number = match.group(1)
                    elif p_name == 'MX64 1X6 CKT (31403)':
                        match = re.search(r"\b(31403\d{4})\b", line_str)
                        if match:
                            ref_number = match.group(1)
                    elif p_name == 'MX123 49 CKT WDC (216744)':
                        match = re.search(r"\b(216744\d{4})\b", line_str)
                        if match:
                            ref_number = match.group(1)
                    elif p_name == 'MXDASH 15+1 RCPT (218342)':
                        match = re.search(r"\b(218342\d{4})\b", line_str)
                        if match:
                            ref_number = match.group(1)
                            
                    if not ref_number:
                        continue
                        
                    tokens = line_str.split()
                    try:
                        idx = tokens.index(ref_number)
                        
                        # 2. Extract abbreviations according to product logic
                        if p_name == 'MX150 2X10 BLD (33482)':
                            remaining = tokens[idx + 1:]
                            grommet_idx = None
                            for i, token in enumerate(remaining):
                                if token in ["GCO", "GCP"]:
                                    grommet_idx = i
                                    break
                            if grommet_idx is not None:
                                housing = " ".join(remaining[:grommet_idx])
                                grommet = remaining[grommet_idx]
                                plr = remaining[grommet_idx + 1]
                                mat = remaining[grommet_idx + 2]
                                pdf_data[ref_number] = {
                                    'HOUSING KEY/COLOR': housing,
                                    'GROMMET CAP': grommet,
                                    'PLR': plr,
                                    'MAT SEAL': mat
                                }
                                
                        elif p_name == 'MX150 2X10 RCPT (33472)':
                            housing = tokens[idx + 1]
                            grommet = tokens[idx + 2]
                            plr = tokens[idx + 3]
                            cpa = ""
                            mat = ""
                            ring = ""
                            if len(tokens) > idx + 4 and tokens[idx + 4] == "CPA":
                                cpa = "CPA"
                                mat = tokens[idx + 5]
                                ring = tokens[idx + 6]
                            else:
                                mat = tokens[idx + 4]
                                ring = tokens[idx + 5]
                                
                            # Normalize housing abbreviation
                            normalized_housing = HOUSING_MAP.get(housing, housing)
                            pdf_data[ref_number] = {
                                'HOUSING KEY/COLOR': normalized_housing,
                                'GROMMET CAP': grommet,
                                'PLR': plr,
                                'CPA': cpa,
                                'MAT SEAL': mat,
                                'RING SEAL': ring
                            }
                            
                        elif p_name == 'CMX 32 CKT (34868)':
                            pdf_data[ref_number] = {
                                'HOUSING': tokens[idx + 1],
                                'TPA-KEY-COLOR': tokens[idx + 2],
                                'R-SEAL': tokens[idx + 3],
                                'RSC': tokens[idx + 4],
                                'M-SEAL': tokens[idx + 5],
                                'LEVER': tokens[idx + 6]
                            }
                            
                        elif p_name == 'MX64 1X6 CKT (31403)':
                            cpa_present = tokens[idx + 1]
                            key = tokens[idx + 2]
                            color = tokens[idx + 3]
                            
                            housing_key = f"{key}-{color}"
                            cpa = "CPA" if cpa_present == "YES" else ""
                            
                            pdf_data[ref_number] = {
                                'CON-HOUSING': housing_key,
                                'TERMINAL HOUSING': 'TML HSG',
                                'TPA': 'TPA',
                                'CPA': cpa,
                                'MAT SEAL': 'M',
                                'RING SEAL': 'R'
                            }
                            
                        elif p_name == 'MX123 49 CKT WDC (216744)':
                            pdf_data[ref_number] = {
                                'CONNECTOR HOUSING': tokens[idx + 1],
                                'GROMMET CAP': tokens[idx + 2],
                                'TPA': tokens[idx + 3],
                                'RIGHT SLIDE': tokens[idx + 4],
                                'LEFT SLIDE': tokens[idx + 5],
                                'MATTE ASSIST LEVER': tokens[idx + 6],
                                'CPA': tokens[idx + 7],
                                'MAT SEAL': tokens[idx + 8],
                                'RING SEAL': tokens[idx + 9]
                            }
                            
                        elif p_name == 'MXDASH 15+1 RCPT (218342)':
                            pdf_data[ref_number] = {
                                'CON-HOUSING': tokens[idx + 1],
                                'TERMINAL HOUSING': tokens[idx + 2],
                                'M-SEAL CAP': tokens[idx + 3],
                                'ISL': tokens[idx + 4],
                                'CPA': tokens[idx + 5],
                                'MAT SEAL': tokens[idx + 6],
                                'RING SEAL': tokens[idx + 7]
                            }
                            
                    except Exception:
                        continue
    except Exception as e:
        print(f"  Warning: Failed to extract BOM from PDF {os.path.basename(pdf_path)}: {e}")
    return pdf_data

def get_sort_key(pair):
    ref, plant = pair
    # Try to convert to int for numeric sort, else leave as string
    try:
        plant_key = int(str(plant).strip())
    except ValueError:
        plant_key = str(plant).strip()
        
    try:
        ref_key = int(str(ref).strip())
    except ValueError:
        ref_key = str(ref).strip()
        
    # Standardize sort keys to compare consistently (int vs string grouping)
    return (
        (0, plant_key) if isinstance(plant_key, int) else (1, str(plant_key)),
        (0, ref_key) if isinstance(ref_key, int) else (1, str(ref_key))
    )

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)
    
    product_names_path = os.path.join(base_dir, "Product Names for UAT1.txt")
    mat_sheet_path = os.path.join(base_dir, "cvp sap tables", "Z9DIR_MAT_SHEET.XLSX")
    template_path = os.path.join(base_dir, "template sheets", "product specific configs", "BOM Template.xlsx")
    family_base_dir = os.path.join(base_dir, "product family")

    print("--- Starting BOM Generation Script (Plant Grouped) ---")

    # 1. Read Product Names from Product Names for UAT1.txt
    if not os.path.exists(product_names_path):
        print(f"Error: Product list not found at {product_names_path}")
        return

    product_names = []
    print("Reading target product list...")
    with open(product_names_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # Remove line numbering prefix
            match = re.match(r'^\d+\.\s*(.*)$', line)
            if match:
                product_names.append(match.group(1).strip())
            else:
                product_names.append(line)
    
    print(f"Loaded {len(product_names)} product(s):")
    for idx, p in enumerate(product_names, start=1):
        print(f"  {idx}. {p}")

    # 2. Build map of Product Folder paths dynamically
    print("\nScanning product family folder structure...")
    product_dirs = {}
    if not os.path.exists(family_base_dir):
        print(f"Error: Family directory not found at {family_base_dir}")
        return

    for root, dirs, _ in os.walk(family_base_dir):
        for d in dirs:
            product_dirs[d] = os.path.join(root, d)

    # 3. Read Material Lookups (Z9DIR_MAT_SHEET.XLSX) using PowerShell bypass
    print("Reading SAP Material Details data...")
    temp_mat_path = os.path.join(base_dir, "cvp sap tables", "temp_Z9DIR_MAT_SHEET_bom.xlsx")
    try:
        if os.path.exists(temp_mat_path):
            os.remove(temp_mat_path)
        cmd = f'Copy-Item -Path "{mat_sheet_path}" -Destination "{temp_mat_path}"'
        subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        df_mat = pd.read_excel(temp_mat_path)
        print(f"Loaded {len(df_mat)} rows from Z9DIR_MAT_SHEET.XLSX successfully.")
    except Exception as e:
        print(f"Error reading SAP Material Details data: {e}")
        return
    finally:
        if os.path.exists(temp_mat_path):
            try:
                os.remove(temp_mat_path)
            except Exception:
                pass

    # Build nested lookups database grouped by Product family
    print("\nIndexing lookups...")
    lookup_db = {}
    for p_name in product_names:
        sap_name = PRODUCT_MAP.get(p_name)
        if not sap_name:
            continue
        sub_df = df_mat[df_mat["Product Name"] == sap_name]
        lookup_db[p_name] = {}
        for _, row in sub_df.iterrows():
            abbrev = str(row.get('Abbreviation', '')).strip().upper()
            if abbrev:
                lookup_db[p_name][abbrev] = {
                    'Description': clean_val(row.get('Description')),
                    'Color': clean_val(row.get('Color')),
                    'Material': clean_val(row.get('Material'))
                }
        print(f"  Indexed {len(lookup_db[p_name])} lookups for '{p_name}'")

    # 4. Populate templates for each product
    print("\nProcessing product BOM configurations...")
    for p_name in product_names:
        # Determine target folder
        target_dir = product_dirs.get(p_name)
        if not target_dir:
            print(f"Warning: No directory found in 'product family/' for '{p_name}'. Skipping.")
            continue

        ref_excel_path = os.path.join(target_dir, "Reference Numbers.xlsx")
        if not os.path.exists(ref_excel_path):
            print(f"Warning: Reference Numbers sheet not found at {ref_excel_path}. Skipping product.")
            continue

        # Load reference numbers and plants using PowerShell lock-bypass
        temp_ref_path = os.path.join(base_dir, "cvp sap tables", f"temp_ref_{p_name.replace(' ', '_').replace('(', '').replace(')', '')}.xlsx")
        try:
            if os.path.exists(temp_ref_path):
                os.remove(temp_ref_path)
            cmd = f'Copy-Item -Path "{ref_excel_path}" -Destination "{temp_ref_path}"'
            subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            df_ref = pd.read_excel(temp_ref_path)
        except Exception as e:
            print(f"Error reading Reference Numbers for '{p_name}': {e}")
            continue
        finally:
            if os.path.exists(temp_ref_path):
                try:
                    os.remove(temp_ref_path)
                except Exception:
                    pass

        # Extract (Reference Number, Plant) pairs
        pairs = []
        for _, row in df_ref.iterrows():
            ref_val = row.get('Reference Number')
            plant_val = row.get('Plant')
            if pd.notna(ref_val) and pd.notna(plant_val):
                pairs.append((format_ref_number(ref_val), format_ref_number(plant_val)))

        if not pairs:
            print(f"Warning: No reference numbers found in Reference Numbers.xlsx for '{p_name}'. Skipping.")
            continue

        # Sort pairs (Plant first, Reference Number second)
        pairs_sorted = sorted(pairs, key=get_sort_key)
        print(f"\nProduct: {p_name} ({len(pairs_sorted)} reference number-plant pair(s) loaded)")

        # Load PDF BOM data for this product
        pdf_path = os.path.join(base_dir, "drawing pdf's", f"{p_name}.pdf")
        pdf_bom_data = parse_pdf_bom_data(pdf_path, p_name)
        print(f"  Loaded {len(pdf_bom_data)} reference number mapping(s) from drawing PDF.")

        output_filepath = os.path.join(target_dir, "BOM.xlsx")

        # Get categories list for this product
        categories = PRODUCT_CATEGORIES.get(p_name, [])
        if not categories:
            print(f"Warning: No categories configured for '{p_name}'. Skipping.")
            continue

        # Build data rows block by block
        data_rows = []
        for ref, plant in pairs_sorted:
            ref_str = str(ref).strip()
            # Try matching with or without leading zeros/spaces
            pdf_ref_key = ref_str
            if pdf_ref_key not in pdf_bom_data:
                # Fallback to matched keys using lstrip or similar
                stripped_key = ref_str.lstrip('0')
                for k in pdf_bom_data:
                    if k.lstrip('0') == stripped_key:
                        pdf_ref_key = k
                        break
            
            ref_data = pdf_bom_data.get(pdf_ref_key, {})
            
            for seq_idx, cat in enumerate(categories, start=1):
                abbrev = ref_data.get(cat, "").strip()
                abbrev_upper = abbrev.upper()
                
                # Check lookup
                description = ""
                color = ""
                material = ""
                
                if abbrev_upper and abbrev_upper in lookup_db.get(p_name, {}):
                    item_lookup = lookup_db[p_name][abbrev_upper]
                    description = item_lookup['Description']
                    color = item_lookup['Color']
                    material = item_lookup['Material']
                
                data_rows.append({
                    'Category': cat,
                    'Abbreviation': abbrev,
                    'Description': description,
                    'Color': color,
                    'Material': material,
                    'Sequence': seq_idx,
                    'PartNumber': seq_idx,
                    'Referencenumber': ref
                })

        # Copy template and write data
        try:
            if os.path.exists(output_filepath):
                try:
                    os.remove(output_filepath)
                except Exception:
                    # Attempt force remove with PowerShell
                    subprocess.run(["powershell", "-Command", f'Remove-Item -Path "{output_filepath}" -Force'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            # Copy template to target location using PowerShell lock-bypass and force overwrite
            cmd = f'Copy-Item -Path "{template_path}" -Destination "{output_filepath}" -Force'
            subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            # Load with openpyxl and populate cells
            wb = openpyxl.load_workbook(output_filepath)
            ws = wb.active
            
            # Clear row 2 if the template has placeholder rows
            if ws.max_row >= 2:
                ws.delete_rows(2, amount=ws.max_row - 1)
            
            # Write data rows starting from row 2
            for r_idx, d_row in enumerate(data_rows, start=2):
                ws.cell(row=r_idx, column=1, value=d_row['Category'])
                ws.cell(row=r_idx, column=2, value=d_row['Abbreviation'])
                ws.cell(row=r_idx, column=3, value=d_row['Description'])
                ws.cell(row=r_idx, column=4, value=d_row['Color'])
                ws.cell(row=r_idx, column=5, value=d_row['Material'])
                ws.cell(row=r_idx, column=6, value=d_row['Sequence'])
                ws.cell(row=r_idx, column=7, value=d_row['PartNumber'])
                ws.cell(row=r_idx, column=8, value=d_row['Referencenumber'])
                
            wb.save(output_filepath)
            wb.close()
            print(f"  Saved BOM sheet successfully to:")
            print(f"  -> {output_filepath}")
        except Exception as e:
            print(f"  Error writing BOM for '{p_name}': {e}")
            print(f"  -> Please make sure '{output_filepath}' is closed in Excel/OneDrive and try again.")
            
    print("\n--- Process Finished ---")

if __name__ == '__main__':
    main()
