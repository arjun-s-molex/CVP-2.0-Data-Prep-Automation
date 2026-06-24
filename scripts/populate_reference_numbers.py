import os
import re
import shutil
import subprocess
import pandas as pd
import openpyxl
import pdfplumber

def clean_val(val):
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val.is_integer():
            val = int(val)
    val_str = str(val).strip()
    if val_str in ("", "-", "nan", "NaN", "None"):
        return ""
    return val_str

def format_ref_number(val):
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return int(val)
    try:
        return int(str(val).strip())
    except ValueError:
        return str(val).strip()

def format_accessory_material(val):
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val.is_integer():
            val = int(val)
    val_str = str(val).strip()
    cleaned_digits = re.sub(r'\D', '', val_str)
    if len(cleaned_digits) >= 9:
        first_part = cleaned_digits[:-4]
        last_part = cleaned_digits[-4:]
        return f"{first_part}-{last_part}"
    elif len(cleaned_digits) > 0:
        return val_str
    else:
        if val_str in ("", "-", "nan", "NaN", "None"):
            return ""
        return val_str

def clean_header(header):
    if not header:
        return ""
    header = re.sub(r'\s+', ' ', header).strip()
    header = header.replace("900", "90°")
    return header

def extract_pdf_accessories(pdf_path):
    if not os.path.exists(pdf_path):
        return {}, False
    
    part_accessories = {}
    has_subheadings = False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    continue
                for table in tables:
                    if len(table) < 3:
                        continue
                    
                    # Find the column index of COMMENT / COMMENTS in Row 0
                    comments_col_idx = -1
                    row0 = table[0]
                    for c_idx, cell in enumerate(row0):
                        if cell:
                            cell_upper = str(cell).upper().strip()
                            if cell_upper in ("COMMENT", "COMMENTS"):
                                comments_col_idx = c_idx
                                break
                    
                    # If comment column is found, the accessory columns are to the right of it
                    if comments_col_idx != -1 and comments_col_idx < len(table[0]) - 1:
                        start_col = comments_col_idx + 1
                        
                        # Subheadings are in Row 2
                        subheading_row_idx = 2
                        subheaders = table[subheading_row_idx][start_col:]
                        cleaned_subheaders = [clean_header(sh) for sh in subheaders]
                        
                        # Check if any subheadings are present
                        if any(cleaned_subheaders):
                            has_subheadings = True
                            
                            # Process data rows
                            for data_r_idx in range(subheading_row_idx + 1, len(table)):
                                d_row = table[data_r_idx]
                                # A valid data row should have assembly part number in col 1
                                if len(d_row) > start_col and d_row[1] and str(d_row[1]).strip().isdigit():
                                    part_num = str(d_row[1]).strip()
                                    attachments = d_row[start_col:]
                                    
                                    parts_list = []
                                    for sh, val in zip(cleaned_subheaders, attachments):
                                        if sh and val and str(val).strip() not in ("N/A", "-", ""):
                                            val_clean = str(val).strip()
                                            cleaned_digits = re.sub(r'\D', '', val_clean)
                                            if len(cleaned_digits) >= 9:
                                                val_formatted = f"{cleaned_digits[:-4]}-{cleaned_digits[-4:]}"
                                            else:
                                                val_formatted = val_clean
                                            parts_list.append(f"{sh}-{val_formatted}")
                                    
                                    if parts_list:
                                        part_key = part_num.lstrip('0')
                                        part_accessories[part_key] = ",".join(parts_list)
    except Exception as e:
        print(f"  Warning: Failed to extract accessories from PDF {os.path.basename(pdf_path)}: {e}")
    return part_accessories, has_subheadings

def main():
    # Define paths (resolved dynamically relative to script location)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)
    
    product_names_path = os.path.join(base_dir, "Product Names for UAT1.txt")
    sap_plant_path = os.path.join(base_dir, "cvp sap tables", "Plant code mapped with Ref number.xlsx")
    sap_accessory_path = os.path.join(base_dir, "cvp sap tables", "Ref PN and Accessory Mapping.xlsx")
    sap_status_path = os.path.join(base_dir, "cvp sap tables", "plant specific material status (2).XLSX")
    template_path = os.path.join(base_dir, "template sheets", "product specific configs", "Reference Numbers Template.xlsx")
    family_base_dir = os.path.join(base_dir, "product family")

    print("--- Starting Reference Numbers Population Script ---")

    # 1. Read Product Names from Product Names for UAT1.txt
    if not os.path.exists(product_names_path):
        print(f"Error: Product list not found at {product_names_path}")
        return

    product_names = []
    print("Reading target product list...")
    with open(product_names_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # Remove line numbering prefix (e.g., "1. CMX 32 CKT (34868)")
            match = re.match(r'^\d+\.\s*(.*)$', line)
            if match:
                product_names.append(match.group(1).strip())
            else:
                product_names.append(line)
    
    print(f"Loaded {len(product_names)} product(s):")
    for idx, p in enumerate(product_names, start=1):
        print(f"  {idx}. {p}")

    # 2. Build map of Product Folder paths dynamically
    print("\nScanning product family folder structure...")
    product_dirs = {}
    if not os.path.exists(family_base_dir):
        print(f"Error: Family directory not found at {family_base_dir}")
        return

    for root, dirs, _ in os.walk(family_base_dir):
        for d in dirs:
            product_dirs[d] = os.path.join(root, d)

    # 3. Read Plant code mapped with Ref number.xlsx (bypassing active file lock)
    print("\nReading SAP Plant Mappings data...")
    temp_plant_path = os.path.join(base_dir, "cvp sap tables", "temp_Plant_code_mapped_Ref.xlsx")
    try:
        if os.path.exists(temp_plant_path):
            os.remove(temp_plant_path)
        cmd = f'Copy-Item -Path "{sap_plant_path}" -Destination "{temp_plant_path}"'
        subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        df_plant = pd.read_excel(temp_plant_path)
        print(f"Loaded {len(df_plant)} rows from Plant code mapped with Ref number.xlsx successfully.")
    except Exception as e:
        print(f"Error reading SAP Plant Mappings data: {e}")
        return
    finally:
        if os.path.exists(temp_plant_path):
            try:
                os.remove(temp_plant_path)
            except Exception:
                pass

    # 4. Read Ref PN and Accessory Mapping.xlsx (bypassing active file lock)
    print("Reading SAP Accessory Mappings data...")
    temp_accessory_path = os.path.join(base_dir, "cvp sap tables", "temp_Ref_PN_Accessory_Ref.xlsx")
    try:
        if os.path.exists(temp_accessory_path):
            os.remove(temp_accessory_path)
        cmd = f'Copy-Item -Path "{sap_accessory_path}" -Destination "{temp_accessory_path}"'
        subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        df_accessory = pd.read_excel(temp_accessory_path)
        print(f"Loaded {len(df_accessory)} rows from Ref PN and Accessory Mapping.xlsx successfully.")
    except Exception as e:
        print(f"  Warning: Could not load Ref PN and Accessory Mapping.xlsx: {e}")
        # Create a dummy DataFrame with all required columns
        df_accessory = pd.DataFrame(columns=[
            'Material', 'BACKSHELL WITH RIBS', 'BACKSHELL WITHOUT RIBS', 
            'ADAPTOR WITH RIBS', 'ADAPTOR WITHOUT RIBS', 
            'DRESS COVER BLACK', 'DRESS COVER BLACK2', 
            'DRESS COVER BLACK3', 'DRESS COVER BLACK4', '90 BACKSHELL',
            '90° ADAPTOR WITH RIBS', '90° ADAPTOR WITHOUT RIBS', '90° BACKSHELL'
        ])
    finally:
        if os.path.exists(temp_accessory_path):
            try:
                os.remove(temp_accessory_path)
            except Exception:
                pass

    # 4.5 Read plant specific material status (2).XLSX (bypassing active file lock)
    print("Reading SAP Material Statuses data...")
    temp_status_path = os.path.join(base_dir, "cvp sap tables", "temp_plant_sp_status_Ref.xlsx")
    try:
        if os.path.exists(temp_status_path):
            os.remove(temp_status_path)
        cmd = f'Copy-Item -Path "{sap_status_path}" -Destination "{temp_status_path}"'
        subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        df_status = pd.read_excel(temp_status_path)
        print(f"Loaded {len(df_status)} rows from plant specific material status (2).XLSX successfully.")
    except Exception as e:
        print(f"Error reading SAP Material Statuses data: {e}")
        return
    finally:
        if os.path.exists(temp_status_path):
            try:
                os.remove(temp_status_path)
            except Exception:
                pass

    # 5. Clean/Rename Accessory Columns to Standardize Names
    rename_cols = {}
    for col in df_accessory.columns:
        col_upper = col.upper()
        if 'BACKSHELL' in col_upper and 'WITHOUT' in col_upper:
            rename_cols[col] = 'BACKSHELL WITHOUT RIBS'
        elif 'BACKSHELL' in col_upper and 'WITH' in col_upper:
            rename_cols[col] = 'BACKSHELL WITH RIBS'
        elif 'ADAPTOR' in col_upper and 'WITHOUT' in col_upper:
            rename_cols[col] = '90° ADAPTOR WITHOUT RIBS'
        elif 'ADAPTOR' in col_upper and 'WITH' in col_upper:
            rename_cols[col] = '90° ADAPTOR WITH RIBS'
        elif '90' in col_upper and 'BACKSHELL' in col_upper:
            rename_cols[col] = '90° BACKSHELL'
        elif 'DRESS COVER' in col_upper:
            rename_cols[col] = col.strip()
    df_accessory = df_accessory.rename(columns=rename_cols)

    # 6. Join the two tables on standardized Material number
    df_plant['join_key'] = df_plant['Material'].astype(str).str.strip().str.lstrip('0')
    df_accessory['join_key'] = df_accessory['Material'].astype(str).str.strip().str.lstrip('0')
    
    # Keep only columns we need from Accessory table to merge cleanly
    acc_columns_to_keep = [
        'join_key',
        'BACKSHELL WITH RIBS',
        'BACKSHELL WITHOUT RIBS',
        '90° ADAPTOR WITH RIBS',
        '90° ADAPTOR WITHOUT RIBS',
        'DRESS COVER BLACK',
        'DRESS COVER BLACK2',
        'DRESS COVER BLACK3',
        'DRESS COVER BLACK4',
        '90° BACKSHELL'
    ]
    df_accessory_subset = df_accessory[[c for c in acc_columns_to_keep if c in df_accessory.columns]]
    merged_df = pd.merge(df_plant, df_accessory_subset, on='join_key', how='left', suffixes=('', '_acc'))

    # Clean keys in merged_df and df_status to merge on Material and Plant Code
    merged_df['join_mat'] = merged_df['Material'].astype(str).str.strip().str.lstrip('0')
    merged_df['join_plant'] = merged_df['Plant Code'].astype(str).str.strip().str.lstrip('0')
    
    df_status['join_mat'] = df_status['Material'].astype(str).str.strip().str.lstrip('0')
    df_status['join_plant'] = df_status['Plant'].astype(str).str.strip().str.lstrip('0')
    
    df_status_subset = df_status[['join_mat', 'join_plant', 'Plant-sp.matl status']]
    merged_df = pd.merge(merged_df, df_status_subset, on=['join_mat', 'join_plant'], how='left')

    # Define columns to extract dependencies
    dep_cols = [
        ('Polarity', 'Polarity'),
        ('Color', 'Color'),
        ('CPA', 'CPA'),
        ('Clipslot', 'Clipslot'),
        ('Grommet', 'Grommet Cap'),
        ('Wire Dress Option', 'Wire Dress Option'),
        ('Wire Output', 'Wire Output'),
        ('Lever Type', 'Lever Type'),
        ('Lever Color', 'Lever Color'),
        ('OEM', 'OEM')
    ]

    # Define accessory columns
    acc_cols = [
        'BACKSHELL WITH RIBS',
        'BACKSHELL WITHOUT RIBS',
        '90° ADAPTOR WITH RIBS',
        '90° ADAPTOR WITHOUT RIBS',
        'DRESS COVER BLACK',
        'DRESS COVER BLACK2',
        'DRESS COVER BLACK3',
        'DRESS COVER BLACK4',
        '90° BACKSHELL'
    ]

    # 7. Map and populate template for each product name
    print("\nProcessing product mappings...")
    for p_name in product_names:
        # Filter rows for this product
        p_df = merged_df[merged_df['Product Name'] == p_name]
        if p_df.empty:
            print(f"Warning: No data found in SAP data for product '{p_name}'. Skipping.")
            continue
        
        print(f"\nProduct: {p_name} ({len(p_df)} row(s) found)")
        
        # Load PDF accessories for this product
        pdf_path = os.path.join(base_dir, "drawing pdf's", f"{p_name}.pdf")
        pdf_acc_dict, has_subheadings = extract_pdf_accessories(pdf_path)
        if has_subheadings:
            print(f"  Accessory columns found in PDF. Loaded {len(pdf_acc_dict)} accessory mapping(s) from PDF (strictly using PDF).")
        else:
            print(f"  No accessory columns/subheadings found in PDF (falling back to Excel).")
        
        # Determine output directory
        target_dir = product_dirs.get(p_name)
        if not target_dir:
            print(f"Warning: No directory found in 'product family/' for '{p_name}'. Skipping.")
            continue
            
        output_filepath = os.path.join(target_dir, "Reference Numbers.xlsx")
        
        # Build target rows
        data_rows = []
        for _, row in p_df.iterrows():
            ref_num = format_ref_number(row.get('Material'))
            plant_code = format_ref_number(row.get('Plant Code'))
            
            # Format dependency string
            dep_parts = []
            for src_col, target_key in dep_cols:
                val = clean_val(row.get(src_col))
                if val:
                    dep_parts.append(f"{target_key}-{val}")
            dep_str = ",".join(dep_parts)
            
            # Format accessory string (strict priority: PDF if subheadings exist, otherwise Excel)
            mat_key = str(ref_num).strip().lstrip('0')
            if has_subheadings:
                acc_str = pdf_acc_dict.get(mat_key)
            else:
                acc_parts = []
                for acc_col in acc_cols:
                    val = row.get(acc_col)
                    formatted_val = format_accessory_material(val)
                    if formatted_val:
                        acc_parts.append(f"{acc_col}-{formatted_val}")
                acc_str = ",".join(acc_parts) if acc_parts else None
            plant_sp_status = clean_val(row.get('Plant-sp.matl status'))

            data_rows.append({
                'Reference Number': ref_num,
                'Product': p_name,
                'Plant': plant_code,
                'Part Status': 'Active',
                'Drawing status': 'Saleable',
                'Plant Sp Status': plant_sp_status,
                'Dependency': dep_str,
                'Accessory': acc_str
            })
            
        # Write to template using openpyxl to keep styles
        try:
            if os.path.exists(output_filepath):
                os.remove(output_filepath)
            
            # Copy template to target location using PowerShell lock-bypass
            cmd = f'Copy-Item -Path "{template_path}" -Destination "{output_filepath}"'
            subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            # Load with openpyxl and populate cells
            wb = openpyxl.load_workbook(output_filepath)
            ws = wb.active
            
            # Clear row 2 if the template has placeholder data
            if ws.max_row >= 2:
                ws.delete_rows(2, amount=ws.max_row - 1)
            
            # Write data rows starting from row 2
            for r_idx, d_row in enumerate(data_rows, start=2):
                ws.cell(row=r_idx, column=1, value=d_row['Reference Number'])
                ws.cell(row=r_idx, column=2, value=d_row['Product'])
                ws.cell(row=r_idx, column=3, value=d_row['Plant'])
                ws.cell(row=r_idx, column=4, value=d_row['Part Status'])
                ws.cell(row=r_idx, column=5, value=d_row['Drawing status'])
                ws.cell(row=r_idx, column=6, value=d_row['Plant Sp Status'])
                ws.cell(row=r_idx, column=7, value=d_row['Dependency'])
                ws.cell(row=r_idx, column=8, value=d_row['Accessory'])
                
            wb.save(output_filepath)
            wb.close()
            print(f"  Saved Reference Numbers sheet successfully to:")
            print(f"  -> {output_filepath}")
        except Exception as e:
            print(f"  Error writing Reference Numbers for '{p_name}': {e}")
            
    print("\n--- Process Finished ---")

if __name__ == '__main__':
    main()
