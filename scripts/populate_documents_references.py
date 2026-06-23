import os
import re
import shutil
import subprocess
import pandas as pd
import openpyxl

def main():
    # Define paths (resolved dynamically relative to script location)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)
    
    product_names_path = os.path.join(base_dir, "Product Names for UAT1.txt")
    sap_table_path = os.path.join(base_dir, "cvp sap tables", "Z9DIR_CVP_PRDLNK.XLSX")
    template_path = os.path.join(base_dir, "template sheets", "product specific configs", "Documents and References Template.xlsx")
    family_base_dir = os.path.join(base_dir, "product family")

    print("--- Starting Documents and References Population Script ---")

    # 1. Read Product Names from Product Names for UAT1.txt
    if not os.path.exists(product_names_path):
        print(f"Error: Product list not found at {product_names_path}")
        return

    product_names = []
    print("Reading target product list...")
    with open(product_names_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # Remove line numbering prefix (e.g., "1. CMX 32 CKT (34868)")
            match = re.match(r'^\d+\.\s*(.*)$', line)
            if match:
                product_names.append(match.group(1).strip())
            else:
                product_names.append(line)
    
    print(f"Loaded {len(product_names)} product(s):")
    for idx, p in enumerate(product_names, start=1):
        print(f"  {idx}. {p}")

    # 2. Build map of Product Folder paths dynamically
    print("\nScanning product family folder structure...")
    product_dirs = {}
    if not os.path.exists(family_base_dir):
        print(f"Error: Family directory not found at {family_base_dir}")
        return

    for root, dirs, _ in os.walk(family_base_dir):
        for d in dirs:
            product_dirs[d] = os.path.join(root, d)

    # 3. Read Z9DIR_CVP_PRDLNK.XLSX file (bypassing active file lock)
    print("\nReading SAP Documents & References data...")
    temp_sap_path = os.path.join(base_dir, "cvp sap tables", "temp_Z9DIR_CVP_PRDLNK.xlsx")
    try:
        # Delete old temp file if exists
        if os.path.exists(temp_sap_path):
            os.remove(temp_sap_path)
        
        # Run powershell copy command
        cmd = f'Copy-Item -Path "{sap_table_path}" -Destination "{temp_sap_path}"'
        subprocess.run(["powershell", "-Command", cmd], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        # Load copy into pandas
        sap_df = pd.read_excel(temp_sap_path)
        print(f"Loaded {len(sap_df)} rows from Z9DIR_CVP_PRDLNK.XLSX successfully.")
    except Exception as e:
        print(f"Error reading SAP Documents & References data: {e}")
        return
    finally:
        # Clean up temp file
        if os.path.exists(temp_sap_path):
            try:
                os.remove(temp_sap_path)
            except Exception:
                pass

    # 4. Map and populate template for each product name
    print("\nProcessing product mappings...")
    for p_name in product_names:
        # Filter rows for this product
        p_df = sap_df[sap_df['Product Name'] == p_name]
        if p_df.empty:
            print(f"Warning: No data found in Z9DIR_CVP_PRDLNK.XLSX for product '{p_name}'. Skipping.")
            continue
        
        print(f"\nProduct: {p_name} ({len(p_df)} row(s) found)")
        
        # Determine output directory
        target_dir = product_dirs.get(p_name)
        if not target_dir:
            print(f"Warning: No directory found in 'product family/' for '{p_name}'. Skipping.")
            continue
            
        output_filepath = os.path.join(target_dir, "Documents and References.xlsx")
        
        # Map values to a list of dictionaries, zero-padding Document Part to 3-digits
        data_rows = []
        for _, row in p_df.iterrows():
            doc_part_val = row.get('Document Part')
            if pd.notna(doc_part_val):
                try:
                    doc_part_str = f"{int(float(doc_part_val)):03d}"
                except (ValueError, TypeError):
                    doc_part_str = str(doc_part_val)
            else:
                doc_part_str = ""

            data_rows.append({
                'Product Name': row.get('Product Name'),
                'Document type': row.get('Document Type'),
                'Document': row.get('Document'),
                'Document Part': doc_part_str
            })
            
        # Write to template using openpyxl to keep styles
        try:
            if os.path.exists(output_filepath):
                os.remove(output_filepath)
            
            # Copy template to target location
            shutil.copy(template_path, output_filepath)
            
            # Load with openpyxl and populate cells
            wb = openpyxl.load_workbook(output_filepath)
            ws = wb.active
            
            # Clear row 2 if the template has placeholder data
            if ws.max_row >= 2:
                ws.delete_rows(2, amount=ws.max_row - 1)
            
            # Write data rows starting from row 2
            for r_idx, d_row in enumerate(data_rows, start=2):
                ws.cell(row=r_idx, column=1, value=d_row['Product Name'])
                ws.cell(row=r_idx, column=2, value=d_row['Document type'])
                ws.cell(row=r_idx, column=3, value=d_row['Document'])
                ws.cell(row=r_idx, column=4, value=d_row['Document Part'])
                
            wb.save(output_filepath)
            wb.close()
            print(f"  Saved Documents and References sheet successfully to:")
            print(f"  -> {output_filepath}")
        except Exception as e:
            print(f"  Error writing Documents and References for '{p_name}': {e}")
            
    print("\n--- Process Finished ---")

if __name__ == '__main__':
    main()
